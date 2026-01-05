"""
Report generation module for video analysis results.
"""

import csv
import json
from datetime import datetime
from pathlib import Path

from .analyzer import VideoAnalysisResult


class ReportGenerator:
    """
    Generate reports from video analysis results.

    Supports Excel (.xlsx), CSV, JSON, and HTML formats.
    """

    def __init__(self, output_dir: str | Path | None = None):
        """
        Initialize the report generator.

        Args:
            output_dir: Directory for output files. Uses current directory if None.
        """
        self.output_dir = Path(output_dir) if output_dir else Path.cwd()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_excel(
        self,
        results: list[VideoAnalysisResult],
        filename: str | None = None,
    ) -> Path:
        """
        Generate an Excel report.

        Args:
            results: List of analysis results.
            filename: Output filename. Auto-generated if None.

        Returns:
            Path to generated file.
        """
        try:
            import openpyxl
        except ImportError as err:
            raise ImportError(
                "openpyxl is required for Excel reports. Install with: pip install openpyxl"
            ) from err

        if filename is None:
            filename = f"video_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = self.output_dir / filename
        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, results)

        # Detailed results sheet
        ws_details = wb.create_sheet("Detailed Results")
        self._create_details_sheet(ws_details, results)

        # Quality scores sheet
        ws_quality = wb.create_sheet("Quality Scores")
        self._create_quality_sheet(ws_quality, results)

        wb.save(filepath)
        return filepath

    def generate_csv(
        self,
        results: list[VideoAnalysisResult],
        filename: str | None = None,
    ) -> Path:
        """
        Generate a CSV report.

        Args:
            results: List of analysis results.
            filename: Output filename. Auto-generated if None.

        Returns:
            Path to generated file.
        """
        if filename is None:
            filename = f"video_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        filepath = self.output_dir / filename

        if not results:
            filepath.touch()
            return filepath

        fieldnames = list(results[0].to_dict().keys())

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for result in results:
                writer.writerow(result.to_dict())

        return filepath

    def generate_json(
        self,
        results: list[VideoAnalysisResult],
        filename: str | None = None,
    ) -> Path:
        """
        Generate a JSON report.

        Args:
            results: List of analysis results.
            filename: Output filename. Auto-generated if None.

        Returns:
            Path to generated file.
        """
        if filename is None:
            filename = f"video_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        filepath = self.output_dir / filename

        data = {
            "generated_at": datetime.now().isoformat(),
            "total_videos": len(results),
            "results": [r.to_dict() for r in results],
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        return filepath

    def generate_html(
        self,
        results: list[VideoAnalysisResult],
        filename: str | None = None,
    ) -> Path:
        """
        Generate an HTML report with visual styling.

        Args:
            results: List of analysis results.
            filename: Output filename. Auto-generated if None.

        Returns:
            Path to generated file.
        """
        if filename is None:
            filename = f"video_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        filepath = self.output_dir / filename

        html = self._generate_html_content(results)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

        return filepath

    def _create_summary_sheet(self, ws, results: list[VideoAnalysisResult]) -> None:
        """Create the summary sheet in Excel."""
        from openpyxl.styles import Font

        # Title
        ws["A1"] = "Video Analysis Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Report info
        ws["A3"] = "Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Total Videos:"
        ws["B4"] = len(results)

        if results:
            # Overall statistics
            ws["A6"] = "Overall Statistics"
            ws["A6"].font = Font(bold=True)

            avg_quality = sum(r.overall_quality_score for r in results) / len(results)
            avg_duplicates = sum(r.duplicate_percentage for r in results) / len(results)
            total_issues = sum(
                len(r.duplicate_frames) + len(r.motion_discontinuities) + len(r.wobble_frames)
                for r in results
            )

            ws["A7"] = "Average Quality Score:"
            ws["B7"] = f"{avg_quality:.1f}"
            ws["A8"] = "Average Duplicate %:"
            ws["B8"] = f"{avg_duplicates:.2f}%"
            ws["A9"] = "Total Issues Found:"
            ws["B9"] = total_issues

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_details_sheet(self, ws, results: list[VideoAnalysisResult]) -> None:
        """Create the detailed results sheet in Excel."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        if not results:
            ws["A1"] = "No results to display"
            return

        headers = list(results[0].to_dict().keys())

        # Header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row, result in enumerate(results, 2):
            data = result.to_dict()
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=data[header])

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_quality_sheet(self, ws, results: list[VideoAnalysisResult]) -> None:
        """Create the quality scores sheet in Excel."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        if not results:
            ws["A1"] = "No results to display"
            return

        headers = ["Filename", "FPS Quality", "Motion Quality", "Duplicate Quality",
                   "Wobble Quality", "Overall Quality"]

        # Header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Color coding for quality scores
        def get_quality_fill(score: float) -> PatternFill:
            if score >= 80:
                return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif score >= 60:
                return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Data rows
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=Path(result.filepath).name)

            scores = [
                result.fps_quality_score,
                result.motion_quality_score,
                result.duplicate_quality_score,
                result.wobble_quality_score,
                result.overall_quality_score,
            ]

            for col, score in enumerate(scores, 2):
                cell = ws.cell(row=row, column=col, value=round(score, 1))
                cell.fill = get_quality_fill(score)

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _generate_html_content(self, results: list[VideoAnalysisResult]) -> str:
        """Generate HTML report content."""
        avg_quality = sum(r.overall_quality_score for r in results) / len(results) if results else 0

        rows_html = ""
        for r in results:
            quality_class = "good" if r.overall_quality_score >= 80 else "warning" if r.overall_quality_score >= 60 else "poor"
            rows_html += f"""
            <tr>
                <td>{Path(r.filepath).name}</td>
                <td>{r.total_frames}</td>
                <td>{r.duration:.2f}s</td>
                <td>{r.fps:.2f}</td>
                <td>{len(r.duplicate_frames)}</td>
                <td>{len(r.motion_discontinuities)}</td>
                <td>{len(r.wobble_frames)}</td>
                <td class="{quality_class}">{r.overall_quality_score:.1f}</td>
            </tr>
            """

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Video Analysis Report</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }}
        h1 {{
            color: #333;
            border-bottom: 2px solid #4472C4;
            padding-bottom: 10px;
        }}
        .summary {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
        }}
        .stat-box {{
            background: #4472C4;
            color: white;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-value {{
            font-size: 2em;
            font-weight: bold;
        }}
        .stat-label {{
            font-size: 0.9em;
            opacity: 0.9;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background: #4472C4;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 12px;
            border-bottom: 1px solid #eee;
        }}
        tr:hover {{
            background: #f9f9f9;
        }}
        .good {{ color: #28a745; font-weight: bold; }}
        .warning {{ color: #ffc107; font-weight: bold; }}
        .poor {{ color: #dc3545; font-weight: bold; }}
        .timestamp {{
            color: #666;
            font-size: 0.9em;
            margin-bottom: 20px;
        }}
    </style>
</head>
<body>
    <h1>📹 Video Analysis Report</h1>
    <p class="timestamp">Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>

    <div class="summary">
        <div class="summary-grid">
            <div class="stat-box">
                <div class="stat-value">{len(results)}</div>
                <div class="stat-label">Videos Analyzed</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{avg_quality:.1f}</div>
                <div class="stat-label">Average Quality Score</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{sum(len(r.duplicate_frames) for r in results)}</div>
                <div class="stat-label">Duplicate Frames</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{sum(len(r.wobble_frames) for r in results)}</div>
                <div class="stat-label">Wobble Issues</div>
            </div>
        </div>
    </div>

    <h2>Detailed Results</h2>
    <table>
        <thead>
            <tr>
                <th>Filename</th>
                <th>Frames</th>
                <th>Duration</th>
                <th>FPS</th>
                <th>Duplicates</th>
                <th>Motion Issues</th>
                <th>Wobble Issues</th>
                <th>Quality Score</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</body>
</html>"""
